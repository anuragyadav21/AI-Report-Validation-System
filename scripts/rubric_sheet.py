"""Populate the reference Rubric worksheet (shared by builder and migration)."""
from __future__ import annotations

from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet


def populate_rubric_sheet(ws: Worksheet) -> None:
    ws.column_dimensions["A"].width = 88
    ws.column_dimensions["B"].width = 28
    title_font = Font(bold=True, size=14)
    h_font = Font(bold=True, size=12)
    body = Alignment(wrap_text=True, vertical="top")
    fill_hdr = PatternFill("solid", fgColor="1F4E79")
    hdr_font = Font(bold=True, color="FFFFFF", size=11)

    r = 1
    ws.cell(r, 1, "Scoring rubric (reference)").font = title_font
    r += 2

    ws.cell(r, 1, "Score anchors (use for every dimension D1–D6)").font = h_font
    r += 1
    anchors = [
        ("Score", "Meaning"),
        ("1", "Very poor"),
        ("2", "Weak"),
        ("3", "Acceptable"),
        ("4", "Good"),
        ("5", "Excellent"),
    ]
    for i, (a, b) in enumerate(anchors):
        row = r + i
        c1 = ws.cell(row=row, column=1, value=a)
        c2 = ws.cell(row=row, column=2, value=b)
        for c in (c1, c2):
            c.font = hdr_font if i == 0 else Font(size=11)
            if i == 0:
                c.fill = fill_hdr
            c.alignment = body
    r += len(anchors) + 2

    dimensions: list[tuple[str, str, str, str]] = [
        (
            "D1 — Groundedness",
            "Unsupported claims, hallucinations, fabricated assumptions.",
            "High: stays close to scenario; labels assumptions; avoids invented details.",
            "Low: fabricates technical specifics; invents evidence; overstates certainty.",
        ),
        (
            "D2 — Structural Completeness",
            "Report organization, section coverage, logical structure.",
            "High: findings, risks, uncertainties, recommendations, conclusion (as appropriate).",
            "Low: disorganized; missing major sections.",
        ),
        (
            "D3 — Technical Depth",
            "Systems-level reasoning, causal analysis, engineering insight.",
            "High: interactions, mechanisms, impacts.",
            "Low: surface-level summary only.",
        ),
        (
            "D4 — Uncertainty Handling",
            "Acknowledgment of ambiguity, confidence calibration, missing data.",
            "High: identifies unknowns; distinguishes facts vs assumptions.",
            "Low: overconfident; ignores uncertainty.",
        ),
        (
            "D5 — Actionability",
            "Usefulness of recommendations.",
            "High: clear mitigations; operational relevance; implementable suggestions.",
            "Low: vague advice; generic statements.",
        ),
        (
            "D6 — Professional Quality",
            "Clarity, technical tone, readability, professionalism.",
            "High: polished engineering communication.",
            "Low: poorly organized or confusing writing.",
        ),
    ]

    for title, measures, high, low in dimensions:
        ws.cell(r, 1, title).font = h_font
        r += 1
        ws.cell(r, 1, f"Measures: {measures}").alignment = body
        r += 1
        ws.cell(r, 1, f"High score: {high}").alignment = body
        r += 1
        ws.cell(r, 1, f"Low score: {low}").alignment = body
        r += 2

    ws.cell(r, 1, "Overall score").font = h_font
    r += 1
    ws.cell(
        r,
        1,
        "Overall = mean of D1 through D6 (computed in column K on Prompt A/B/C sheets).",
    ).alignment = body
