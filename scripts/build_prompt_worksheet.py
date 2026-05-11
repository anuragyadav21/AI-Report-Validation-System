"""Prompt workbook: create a fresh template, or sync rubric layout on an existing file.

Default behavior preserves all data in columns A–D (including LLM responses). Only
``--force-rebuild`` writes a brand-new workbook and wipes the file.
"""
from __future__ import annotations

import argparse
import sys
from collections.abc import Callable
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

from experiment_constants import (
    COL_FIRST_SCORE,
    COL_LAST_SCORE,
    COL_LLM_RESPONSE,
    COL_NOTES,
    COL_OVERALL,
    COL_PROMPT_SCENARIO,
    COL_SCENARIO,
    COL_SERIAL,
    FIRST_DATA_ROW,
    LAST_TEMPLATE_ROW,
    PROMPT_SHEETS,
    RUBRIC_HEADERS,
)
from rubric_sheet import populate_rubric_sheet
from workbook_layout import apply_score_validation, ensure_rubric_columns_in_workbook

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "prompt-worksheet.xlsx"

SCENARIO_HEADER = "Scenario"
SERIAL_HEADER = "S. No."
COMPOSED_PROMPT_HEADER = "Prompt + Scenario"
LLM_RESPONSE_HEADER = "LLM response"

# Scenario task text for column B (same report type / framework; different situations).
SCENARIOS: list[str] = [
    """Autonomous Drone Navigation Failure

An autonomous delivery drone experienced unstable flight behavior during a routine delivery mission after temporary GPS signal degradation. Operators observed sudden altitude fluctuations and route deviation before the drone initiated an emergency landing.

Generate a systems engineering assessment report analyzing possible causes, operational risks, uncertainties, and recommended mitigation actions.""",
    """Railway Signaling Synchronization Issue

A railway signaling control system experienced a software synchronization issue that caused temporary route assignment conflicts between adjacent track sections. No collision occurred, but several trains experienced emergency braking events.

Generate a systems engineering assessment report evaluating system risks, contributing factors, uncertainties, and corrective actions.""",
    """Smart Grid Load Balancing Instability

A smart electrical grid experienced unexpected regional voltage fluctuations during peak demand hours after an automated load-balancing algorithm redistributed power unevenly across substations.

Generate a systems engineering assessment report analyzing technical risks, system vulnerabilities, uncertainties, and mitigation strategies.""",
    """Hospital Patient Monitoring Delay

An ICU patient monitoring system generated delayed alerts due to intermittent sensor communication latency. Clinical staff reported several cases where abnormal patient conditions were identified manually before automated alerts were triggered.

Generate a systems engineering assessment report assessing operational risks, technical concerns, uncertainties, and recommended improvements.""",
    """Manufacturing Robot Control Failure

An industrial robotic assembly arm stopped responding during automated manufacturing operations, resulting in temporary production shutdown and manual intervention by operators.

Generate a systems engineering assessment report identifying potential causes, operational impacts, uncertainties, and mitigation recommendations.""",
    """Aircraft Autopilot Disengagement Anomaly

During simulation testing, an aircraft autopilot system unexpectedly disengaged during descent operations after receiving inconsistent sensor input from redundant altitude sensors.

Generate a systems engineering assessment report evaluating safety implications, possible failure modes, uncertainties, and recommended corrective actions.""",
    """Cybersecurity Monitoring Detection Failure

A network intrusion detection system failed to flag abnormal outbound traffic associated with unauthorized data transfer activity. Security analysts later identified the anomaly during a manual review process.

Generate a systems engineering assessment report analyzing detection gaps, operational risks, uncertainties, and recommended mitigations.""",
    """Satellite Telemetry Communication Loss

A satellite communication platform experienced intermittent telemetry loss during orbital adjustment maneuvers, reducing visibility into subsystem health and navigation performance.

Generate a systems engineering assessment report examining possible technical causes, mission risks, uncertainties, and mitigation strategies.""",
    """Autonomous Vehicle Sensor Conflict

An autonomous vehicle platform reported conflicting object detection outputs between LiDAR and camera-based perception systems during low-visibility driving conditions.

Generate a systems engineering assessment report assessing technical risks, uncertainty factors, system limitations, and recommended corrective actions.""",
    """Industrial Temperature Control Instability

A manufacturing plant’s automated temperature regulation controller produced unstable thermal behavior during high-load operations, resulting in inconsistent product quality and temporary production delays.

Generate a systems engineering assessment report analyzing operational impacts, possible root causes, uncertainties, and recommended mitigation actions.""",
]


def compose_prompt_a(scenario: str) -> str:
    return f"""Prompt A + Scenario
Write a report about this incident.

{scenario}"""


def compose_prompt_b(scenario: str) -> str:
    return f"""Prompt B + Scenario
Generate a technical systems engineering report.

Include:
- operational impacts
- risks
- mitigation strategies
- limitations

{scenario}"""


def compose_prompt_c(scenario: str) -> str:
    return f"""Prompt C + Scenario
Role: You are a senior IV&V systems engineer.

Generate a formal assessment report.

Requirements:
- identify failure modes
- discuss uncertainty
- distinguish facts from inference
- provide traceability-oriented reasoning
- avoid unsupported claims
- include actionable mitigation recommendations

{scenario}"""


SHEETS: list[tuple[str, Callable[[str], str]]] = [
    ("Prompt A", compose_prompt_a),
    ("Prompt B", compose_prompt_b),
    ("Prompt C", compose_prompt_c),
]


def apply_sheet_layout(ws: Worksheet, compose: Callable[[str], str]) -> None:
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    wrap = Alignment(wrap_text=True, vertical="top")
    center_header = Alignment(
        wrap_text=True, vertical="center", horizontal="center"
    )

    base_headers = [
        SERIAL_HEADER,
        SCENARIO_HEADER,
        COMPOSED_PROMPT_HEADER,
        LLM_RESPONSE_HEADER,
    ]
    all_headers = base_headers + RUBRIC_HEADERS
    for col, text in enumerate(all_headers, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_header
        cell.border = border

    for i, scenario in enumerate(SCENARIOS):
        r = i + 2
        serial_cell = ws.cell(row=r, column=COL_SERIAL, value=i + 1)
        scenario_cell = ws.cell(row=r, column=COL_SCENARIO, value=scenario)
        composed_cell = ws.cell(row=r, column=COL_PROMPT_SCENARIO, value=compose(scenario))
        serial_cell.alignment = Alignment(
            wrap_text=True, vertical="top", horizontal="center"
        )
        serial_cell.border = border
        scenario_cell.alignment = wrap
        scenario_cell.border = border
        composed_cell.alignment = wrap
        composed_cell.border = border

        llm_cell = ws.cell(row=r, column=COL_LLM_RESPONSE, value=None)
        llm_cell.alignment = wrap
        llm_cell.border = border

        for c in range(COL_FIRST_SCORE, COL_LAST_SCORE + 1):
            cell = ws.cell(row=r, column=c, value=None)
            cell.alignment = Alignment(wrap_text=True, vertical="top", horizontal="center")
            cell.border = border

        overall_cell = ws.cell(
            row=r,
            column=COL_OVERALL,
            value=f"=IFERROR(AVERAGE(E{r}:J{r}),\"\")",
        )
        overall_cell.alignment = Alignment(
            wrap_text=True, vertical="top", horizontal="center"
        )
        overall_cell.border = border
        overall_cell.number_format = "0.00"

        notes_cell = ws.cell(row=r, column=COL_NOTES, value=None)
        notes_cell.alignment = wrap
        notes_cell.border = border

    for r in range(2 + len(SCENARIOS), LAST_TEMPLATE_ROW + 1):
        for c in range(COL_SERIAL, COL_NOTES + 1):
            cell = ws.cell(row=r, column=c, value=None)
            cell.alignment = wrap
            cell.border = border
        ws.cell(
            row=r,
            column=COL_OVERALL,
            value=f"=IFERROR(AVERAGE(E{r}:J{r}),\"\")",
        ).number_format = "0.00"
        ws.cell(row=r, column=COL_OVERALL).alignment = Alignment(
            wrap_text=True, vertical="top", horizontal="center"
        )
        ws.cell(row=r, column=COL_OVERALL).border = border
        for c in range(COL_FIRST_SCORE, COL_LAST_SCORE + 1):
            ws.cell(row=r, column=c).alignment = Alignment(
                wrap_text=True, vertical="top", horizontal="center"
            )

    apply_score_validation(ws)

    ws.freeze_panes = "E2"
    ws.row_dimensions[1].height = 40
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 52
    ws.column_dimensions["C"].width = 62
    ws.column_dimensions["D"].width = 48
    for col_letter in ("E", "F", "G", "H", "I", "J"):
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 36


def _create_fresh_workbook(path: Path) -> None:
    wb = Workbook()
    first_name, first_compose = SHEETS[0]
    ws0 = wb.active
    ws0.title = first_name
    apply_sheet_layout(ws0, first_compose)

    for tab_name, compose in SHEETS[1:]:
        ws = wb.create_sheet(title=tab_name)
        apply_sheet_layout(ws, compose)

    rub = wb.create_sheet(title="Rubric")
    populate_rubric_sheet(rub)

    wb.save(path)
    print(f"Wrote fresh template: {path}")


def main() -> None:
    ap = argparse.ArgumentParser(
        description=__doc__
        + " By default, an existing workbook is only synced (rubric columns / Rubric sheet); "
        "column D (LLM responses) is never cleared.",
    )
    ap.add_argument(
        "--output",
        type=Path,
        default=OUT,
        help="Workbook path (default: project prompt-worksheet.xlsx)",
    )
    ap.add_argument(
        "--force-rebuild",
        action="store_true",
        help="Overwrite output with a NEW blank workbook (deletes existing LLM and manual data).",
    )
    args = ap.parse_args()
    out: Path = args.output

    if not out.exists():
        _create_fresh_workbook(out)
        return

    if not args.force_rebuild:
        wb = load_workbook(out)
        missing = [n for n in PROMPT_SHEETS if n not in wb.sheetnames]
        if missing:
            print(
                f"Workbook exists but is missing sheets {missing}. "
                "Use --force-rebuild to create a full template, or fix the file.",
                file=sys.stderr,
            )
            raise SystemExit(1)
        try:
            changed = ensure_rubric_columns_in_workbook(wb)
        except ValueError:
            raise SystemExit(1) from None
        if changed:
            wb.save(out)
            print(f"Synced layout (preserved columns A–D): {out}")
        else:
            print(f"No layout changes needed (left file untouched): {out}")
        return

    print(
        "WARNING: --force-rebuild replaces the entire workbook (all data in file is lost).",
        file=sys.stderr,
    )
    _create_fresh_workbook(out)


if __name__ == "__main__":
    main()
