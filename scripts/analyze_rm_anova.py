#!/usr/bin/env python3
"""Repeated-measures ANOVA on Overall scores (Prompt A vs B vs C) from the experiment workbook.

Assumptions (classical one-way RM ANOVA): multivariate normality of difference scores;
sphericity of differences between prompt levels (Pingouin uses Mauchly / Greenhouse–Geiser
when correction='auto'). If cell counts are small or scores are non-normal, interpret
with caution; a non-parametric fallback is the Friedman test on the same wide matrix.
"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

_ROOT = Path(__file__).resolve().parents[1]
_mpl = _ROOT / ".mpl_cache"
_mpl.mkdir(exist_ok=True)
os.environ.setdefault("MPLCONFIGDIR", str(_mpl))

import numpy as np
import pandas as pd
import pingouin as pg
from openpyxl import load_workbook
from scipy.stats import friedmanchisquare

from experiment_constants import (
    COL_FIRST_SCORE,
    COL_LAST_SCORE,
    COL_OVERALL,
    COL_SERIAL,
    FIRST_DATA_ROW,
    LAST_DATA_ROW,
    PROMPT_SHEETS,
)

ROOT = _ROOT
DEFAULT_XLSX = ROOT / "prompt-worksheet.xlsx"

SHEET_PROMPT = {"Prompt A": "A", "Prompt B": "B", "Prompt C": "C"}


def _cell_float(ws, row: int, col: int) -> float | None:
    v = ws.cell(row=row, column=col).value
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        return float(v)
    return None


def _row_overall(ws, row: int) -> float | None:
    """Prefer column K (Overall) when it is numeric — matches the sheet's final score.

    Fall back to mean(D1–D6) only if K has no stored number (e.g. blank or formula not cached
    in data_only mode). Previously mean(E:J) was always used when present, which ignored
    overwritten Overall values in K (e.g. after synthetic edits).
    """
    k = _cell_float(ws, row, COL_OVERALL)
    if k is not None:
        return k
    vals: list[float] = []
    for c in range(COL_FIRST_SCORE, COL_LAST_SCORE + 1):
        x = _cell_float(ws, row, c)
        if x is not None:
            vals.append(x)
    if not vals:
        return None
    return float(np.mean(vals))


def build_long_frame(path: Path) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True)
    rows: list[dict] = []
    for sheet in PROMPT_SHEETS:
        if sheet not in wb.sheetnames:
            raise ValueError(f"Missing sheet {sheet!r}")
        ws = wb[sheet]
        prompt = SHEET_PROMPT[sheet]
        for r in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            sid = ws.cell(row=r, column=COL_SERIAL).value
            if sid is None:
                continue
            rec: dict = {
                "ScenarioID": sid,
                "Prompt": prompt,
                "Overall": _row_overall(ws, r),
            }
            for j, c in enumerate(range(COL_FIRST_SCORE, COL_LAST_SCORE + 1), start=1):
                rec[f"D{j}"] = _cell_float(ws, r, c)
            rows.append(rec)
    return pd.DataFrame(rows)


def partial_eta_sq(ss_effect: float, ss_error: float) -> float | None:
    if ss_effect + ss_error <= 0:
        return None
    return ss_effect / (ss_effect + ss_error)


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", type=Path, default=DEFAULT_XLSX)
    ap.add_argument(
        "--csv-out",
        type=Path,
        default=None,
        help="Write long-format table (ScenarioID, Prompt, D1–D6, Overall) to this path",
    )
    args = ap.parse_args()

    if not args.workbook.is_file():
        print(f"Workbook not found: {args.workbook}", file=sys.stderr)
        return 1

    try:
        df = build_long_frame(args.workbook)
    except ValueError as e:
        print(e, file=sys.stderr)
        return 1

    if args.csv_out:
        df.to_csv(args.csv_out, index=False)
        print(f"Wrote {args.csv_out}")

    need = {"ScenarioID", "Prompt", "Overall"}
    if not need.issubset(df.columns):
        print("Unexpected frame columns", file=sys.stderr)
        return 1

    comp = df.dropna(subset=["Overall"]).copy()
    if comp.empty:
        print("No numeric Overall scores (enter D1–D6 or computed Overall in Excel).", file=sys.stderr)
        return 1

    wide = comp.pivot_table(
        index="ScenarioID", columns="Prompt", values="Overall", aggfunc="first"
    )
    if not {"A", "B", "C"}.issubset(wide.columns):
        print("Need all three prompt levels in the workbook.", file=sys.stderr)
        return 1

    wide_complete = wide.dropna(subset=["A", "B", "C"], how="any")
    if len(wide_complete) < 2:
        print(
            f"Need at least 2 complete scenarios (non-missing A,B,C); got {len(wide_complete)}.",
            file=sys.stderr,
        )
        return 1

    long = wide_complete.reset_index().melt(
        id_vars="ScenarioID",
        var_name="Prompt",
        value_name="Overall",
    )

    print("=== Descriptive: Overall by Prompt (complete cases) ===")
    print(long.groupby("Prompt")["Overall"].agg(["count", "mean", "std"]).round(4))
    print()

    aov = pg.rm_anova(
        data=long,
        dv="Overall",
        within="Prompt",
        subject="ScenarioID",
        detailed=True,
        correction="auto",
    )
    print("=== Repeated-measures ANOVA (Overall ~ Prompt) ===")
    print(aov.to_string(index=False))
    print()

    prompt_row = aov[aov["Source"] == "Prompt"]
    if not prompt_row.empty:
        ss_p = float(prompt_row["SS"].iloc[0])
        err_rows = aov[aov["Source"] == "Error"]
        if not err_rows.empty:
            ss_e = float(err_rows["SS"].iloc[0])
            np2 = partial_eta_sq(ss_p, ss_e)
            if np2 is not None:
                print(f"Partial eta-squared (Prompt): {np2:.4f}")
                print("  (conventional cutoffs: ~0.01 small, ~0.06 medium, ~0.14 large)")
        print()

    post = pg.pairwise_tests(
        data=long,
        dv="Overall",
        within="Prompt",
        subject="ScenarioID",
        padjust="holm",
        effsize="hedges",
    )
    print("=== Pairwise comparisons (paired tests, Holm-adjusted p) ===")
    print(post.to_string(index=False))
    print()

    fa, fb, fc = (
        wide_complete["A"].to_numpy(),
        wide_complete["B"].to_numpy(),
        wide_complete["C"].to_numpy(),
    )
    try:
        fr_stat, fr_p = friedmanchisquare(fa, fb, fc)
        print("=== Non-parametric fallback: Friedman test (same complete cases) ===")
        print(f"  chi2({len(fa)} blocks) = {fr_stat:.4f}, p = {fr_p:.4g}")
    except ValueError as e:
        print(f"Friedman test skipped: {e}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
