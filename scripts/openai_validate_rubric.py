#!/usr/bin/env python3
"""PART 2 — Validator / judge: score each LLM report (D) vs scenario (B) using the rubric; write E–J, L.

Uses OpenAI Chat Completions with strict JSON output. Preserves column D.
By default skips rows where E is already filled; use --overwrite-scores to replace.
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import time
from pathlib import Path
from typing import Any

from openai import OpenAI
from openpyxl import load_workbook

from experiment_constants import (
    COL_FIRST_SCORE,
    COL_LAST_SCORE,
    COL_LLM_RESPONSE,
    COL_NOTES,
    COL_SCENARIO,
    PROMPT_SHEETS,
    FIRST_DATA_ROW,
    LAST_DATA_ROW,
)
from openai_env import get_openai_api_key

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "prompt-worksheet.xlsx"

SYSTEM_MESSAGE = """You are an expert systems engineering evaluator.

Your task is to evaluate AI-generated systems engineering reports using a structured scoring rubric.

Return ONLY valid JSON (a single JSON object, no markdown, no prose before or after)."""

SCORING_RUBRIC_BLOCK = """
D1 Groundedness

Measures unsupported claims and hallucinations.
1 = heavily fabricated
5 = fully grounded in scenario

D2 Structural Completeness

Measures report organization and section coverage.
1 = disorganized
5 = comprehensive and complete

D3 Technical Depth

Measures systems-level reasoning and analytical insight.
1 = superficial
5 = highly analytical

D4 Uncertainty Handling

Measures acknowledgment of ambiguity and missing data.
1 = overconfident
5 = excellent uncertainty calibration

D5 Actionability

Measures usefulness of recommendations.
1 = vague
5 = highly actionable

D6 Professional Quality

Measures clarity and professionalism.
1 = poor communication
5 = polished technical writing
"""

JSON_SCHEMA_INSTRUCTIONS = """
Return ONLY valid JSON using this exact structure (all six scores are integers 1–5):

{
  "D1_groundedness": <integer>,
  "D2_structure": <integer>,
  "D3_technical_depth": <integer>,
  "D4_uncertainty": <integer>,
  "D5_actionability": <integer>,
  "D6_professional_quality": <integer>,
  "overall_mean": <float>,
  "comments": "<brief explanation string>"
}
"""

EXPECTED_KEYS = (
    "D1_groundedness",
    "D2_structure",
    "D3_technical_depth",
    "D4_uncertainty",
    "D5_actionability",
    "D6_professional_quality",
    "overall_mean",
    "comments",
)

KEY_TO_COL = {
    "D1_groundedness": 5,
    "D2_structure": 6,
    "D3_technical_depth": 7,
    "D4_uncertainty": 8,
    "D5_actionability": 9,
    "D6_professional_quality": 10,
}


def strip_json_fences(text: str) -> str:
    t = text.strip()
    if t.startswith("```"):
        t = re.sub(r"^```(?:json)?\s*", "", t, flags=re.IGNORECASE)
        t = re.sub(r"\s*```\s*$", "", t)
    return t.strip()


def parse_validator_json(raw: str) -> dict[str, Any]:
    cleaned = strip_json_fences(raw)
    return json.loads(cleaned)


def validate_and_normalize_scores(data: dict[str, Any]) -> dict[str, Any]:
    for k in EXPECTED_KEYS:
        if k not in data:
            raise ValueError(f"missing key {k!r}")
    out: dict[str, Any] = {}
    for k in KEY_TO_COL:
        v = data[k]
        if isinstance(v, bool):
            raise ValueError(f"{k} must be integer, got bool")
        if isinstance(v, float) and abs(v - round(v)) < 1e-9:
            v = int(round(v))
        if not isinstance(v, int):
            raise ValueError(f"{k} must be integer, got {type(v).__name__}")
        if v < 1 or v > 5:
            raise ValueError(f"{k} must be 1–5, got {v}")
        out[k] = v
    om = data["overall_mean"]
    if not isinstance(om, (int, float)) or isinstance(om, bool):
        raise ValueError("overall_mean must be a number")
    out["overall_mean"] = float(om)
    cm = data["comments"]
    if not isinstance(cm, str):
        raise ValueError("comments must be a string")
    out["comments"] = cm.strip()[:5000]
    return out


def build_user_message(scenario: str, report: str) -> str:
    return f"""Evaluate the following AI-generated systems engineering assessment report.

SCENARIO:
{scenario}

GENERATED REPORT:
{report}

SCORING RUBRIC:
{SCORING_RUBRIC_BLOCK}

{JSON_SCHEMA_INSTRUCTIONS}
"""


def row_has_scores(ws, row: int) -> bool:
    v = ws.cell(row=row, column=COL_FIRST_SCORE).value
    return v is not None and str(v).strip() != ""


def call_validator(
    client: OpenAI,
    *,
    model: str,
    user_content: str,
    max_tokens: int,
) -> str:
    resp = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": SYSTEM_MESSAGE},
            {"role": "user", "content": user_content},
        ],
        max_tokens=max_tokens,
        temperature=0.1,
        response_format={"type": "json_object"},
    )
    choice = resp.choices[0]
    content = choice.message.content
    if not content:
        return ""
    return content


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--workbook", type=Path, default=DEFAULT_XLSX)
    ap.add_argument("--model", default="gpt-4o-mini")
    ap.add_argument("--max-tokens", type=int, default=2048)
    ap.add_argument("--timeout", type=float, default=120.0)
    ap.add_argument(
        "--overwrite-scores",
        action="store_true",
        help="Write E–J / L even when scores already exist (default: skip)",
    )
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    api_key = get_openai_api_key(required=not args.dry_run)

    path = args.workbook
    if not path.is_file():
        print(f"Workbook not found: {path}", file=sys.stderr)
        return 1

    client = OpenAI(api_key=api_key, timeout=args.timeout) if not args.dry_run else None

    wb = load_workbook(path)
    for name in PROMPT_SHEETS:
        if name not in wb.sheetnames:
            print(f"Missing sheet {name!r}", file=sys.stderr)
            return 1

    total = 0
    for sheet_name in PROMPT_SHEETS:
        ws = wb[sheet_name]
        for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            if not args.overwrite_scores and row_has_scores(ws, row):
                print(f"{sheet_name} row {row}: skip (scores present; use --overwrite-scores)")
                continue

            scenario = ws.cell(row=row, column=COL_SCENARIO).value
            report = ws.cell(row=row, column=COL_LLM_RESPONSE).value
            if not report or not str(report).strip():
                print(f"{sheet_name} row {row}: skip (no LLM response in column D)")
                continue
            scenario = str(scenario or "").strip()
            report = str(report).strip()

            user_msg = build_user_message(scenario, report)
            total += 1
            print(f"[{total}] {sheet_name} row {row} -> validator ({args.model}) ...", flush=True)

            if args.dry_run:
                print(f"  dry-run, user message length {len(user_msg)} chars")
                continue

            assert client is not None
            raw = None
            for attempt in range(3):
                try:
                    raw = call_validator(
                        client,
                        model=args.model,
                        user_content=user_msg,
                        max_tokens=args.max_tokens,
                    )
                    data = parse_validator_json(raw)
                    norm = validate_and_normalize_scores(data)
                    break
                except Exception as e:
                    wait = 2**attempt
                    print(f"  attempt {attempt + 1} failed: {e}; retry in {wait}s", flush=True)
                    if attempt == 2:
                        print(f"  giving up on {sheet_name} row {row}", file=sys.stderr)
                        return 1
                    time.sleep(wait)
            else:
                return 1

            for key, col in KEY_TO_COL.items():
                ws.cell(row=row, column=col, value=norm[key])

            manual_mean = sum(norm[k] for k in KEY_TO_COL) / 6.0
            om = norm["overall_mean"]
            note = norm["comments"]
            if abs(om - manual_mean) > 0.01:
                note = (
                    f"{note}\n\n[validator overall_mean={om}; mean(D1–D6)={manual_mean:.4f}]"
                )
            ws.cell(row=row, column=COL_NOTES, value=note)

            wb.save(path)
            print(
                f"  wrote D1–D6 + Notes; column K Overall is formula from scores ({len(note)} chars notes)",
                flush=True,
            )

    if args.dry_run:
        print(f"Dry-run: would validate {total} rows.")
    else:
        print(f"Done. Processed {total} rows in {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
