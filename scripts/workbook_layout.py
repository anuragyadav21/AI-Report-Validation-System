"""Shared layout helpers: rubric columns (E–L), validation, Rubric sheet — without touching A–D."""
from __future__ import annotations

import sys

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from experiment_constants import (
    COL_FIRST_SCORE,
    COL_LAST_SCORE,
    COL_NOTES,
    COL_OVERALL,
    FIRST_DATA_ROW,
    LAST_TEMPLATE_ROW,
    PROMPT_SHEETS,
    RUBRIC_HEADERS,
)
from rubric_sheet import populate_rubric_sheet


def apply_score_validation(ws: Worksheet) -> None:
    dv = DataValidation(
        type="whole",
        operator="between",
        formula1=1,
        formula2=5,
        allow_blank=True,
    )
    dv.error = "Enter a whole number from 1 to 5 (or leave blank)."
    dv.errorTitle = "Invalid score"
    ws.add_data_validation(dv)
    dv.add(f"E{FIRST_DATA_ROW}:J{LAST_TEMPLATE_ROW}")


def style_rubric_block_on_sheet(ws: Worksheet) -> None:
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    center_header = Alignment(
        wrap_text=True, vertical="center", horizontal="center"
    )
    wrap = Alignment(wrap_text=True, vertical="top")
    center_score = Alignment(wrap_text=True, vertical="top", horizontal="center")

    for col, text in enumerate(RUBRIC_HEADERS, start=COL_FIRST_SCORE):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_header
        cell.border = border

    for r in range(FIRST_DATA_ROW, LAST_TEMPLATE_ROW + 1):
        for c in range(COL_FIRST_SCORE, COL_LAST_SCORE + 1):
            cell = ws.cell(row=r, column=c)
            cell.alignment = center_score
            cell.border = border

        oc = ws.cell(
            row=r,
            column=COL_OVERALL,
            value=f"=IFERROR(AVERAGE(E{r}:J{r}),\"\")",
        )
        oc.number_format = "0.00"
        oc.alignment = center_score
        oc.border = border

        nc = ws.cell(row=r, column=COL_NOTES)
        nc.alignment = wrap
        nc.border = border

    ws.freeze_panes = "E2"
    for col_letter in ("E", "F", "G", "H", "I", "J"):
        ws.column_dimensions[col_letter].width = 14
    ws.column_dimensions["K"].width = 10
    ws.column_dimensions["L"].width = 36


def ensure_rubric_columns_in_workbook(wb: Workbook) -> bool:
    """Insert rubric columns / Rubric sheet if missing. Does not modify columns A–D."""
    changed = False
    for name in PROMPT_SHEETS:
        if name not in wb.sheetnames:
            print(f"Missing sheet {name!r}", file=sys.stderr)
            raise ValueError(f"missing sheet {name}")
        ws = wb[name]
        probe = ws.cell(row=1, column=COL_FIRST_SCORE).value
        if probe and "D1" in str(probe):
            print(f"{name}: rubric columns already present — skipping insert.")
            continue

        ws.insert_cols(COL_FIRST_SCORE, amount=len(RUBRIC_HEADERS))
        style_rubric_block_on_sheet(ws)
        apply_score_validation(ws)
        changed = True

    if "Rubric" not in wb.sheetnames:
        rub = wb.create_sheet(title="Rubric")
        populate_rubric_sheet(rub)
        print("Added sheet 'Rubric'.")
        changed = True
    else:
        print("Sheet 'Rubric' already exists — leaving unchanged.")

    return changed
