#!/usr/bin/env python3
"""Copy the workbook; on selected prompt sheets optionally down-randomize D1–D6 and/or replace Overall (K).

1) Optional: subtract 2 points total across D1–D6 per row (random dimensions, scores stay 1–5).
2) Optional: replace column K with a uniform random **final** score in a sheet-specific range (removes
   the Excel formula on K).

Default: **Prompt A** only — random Overall K in [3.9, 4.3].
With **--prompt-b**, also process **Prompt B** — random Overall K in [4.2, 4.67] by default.

Other sheets are unchanged. Original file is never modified (copy first).

Output defaults to prompt-worksheet_promptA_minus2.xlsx next to the input.
"""
from __future__ import annotations

import argparse
import random
import shutil
import sys
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from experiment_constants import (
    COL_FIRST_SCORE,
    COL_LAST_SCORE,
    COL_NOTES,
    COL_OVERALL,
    FIRST_DATA_ROW,
    LAST_DATA_ROW,
)

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_IN = ROOT / "prompt-worksheet.xlsx"
DEFAULT_OUT = ROOT / "prompt-worksheet_promptA_minus2.xlsx"

NOTE_OVERALL = " [Overall K: synthetic uniform random; formula replaced]"


def _read_scores(ws: Worksheet, row: int) -> list[int] | None:
    vals: list[int] = []
    for c in range(COL_FIRST_SCORE, COL_LAST_SCORE + 1):
        v = ws.cell(row=row, column=c).value
        if v is None or v == "":
            return None
        if isinstance(v, float) and abs(v - round(v)) < 1e-9:
            v = int(round(v))
        if not isinstance(v, int) or isinstance(v, bool):
            return None
        if v < 1 or v > 5:
            return None
        vals.append(v)
    return vals


def _apply_minus_two_random(scores: list[int]) -> list[int]:
    s = list(scores)
    remaining = 2
    while remaining > 0:
        eligible = [i for i in range(6) if s[i] > 1]
        if not eligible:
            break
        i = random.choice(eligible)
        s[i] -= 1
        remaining -= 1
    return s


def _append_note(ws: Worksheet, row: int, fragment: str) -> None:
    cell = ws.cell(row=row, column=COL_NOTES)
    prev = cell.value
    if prev is None or str(prev).strip() == "":
        cell.value = fragment.strip()
    else:
        cell.value = str(prev).rstrip() + fragment


def process_sheet(
    ws: Worksheet,
    sheet_name: str,
    *,
    skip_downscore: bool,
    no_random_overall: bool,
    overall_min: float,
    overall_max: float,
) -> tuple[int, int]:
    """Returns (downscore_row_count, random_overall_row_count)."""
    note_down = (
        f" [{sheet_name} clone: −2 pts total distributed randomly across D1–D6]"
    )
    down_changed = 0
    for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
        old = _read_scores(ws, row)
        if old is None:
            print(f"{sheet_name} row {row}: skip downscore (missing or non-integer D1–D6)")
            continue
        if skip_downscore:
            print(f"{sheet_name} row {row}: downscore skipped (--skip-downscore)")
            continue
        new = _apply_minus_two_random(old)
        if new == old:
            print(
                f"{sheet_name} row {row}: skip downscore (could not subtract 2; all at floor 1)"
            )
            continue
        for c, val in enumerate(new, start=COL_FIRST_SCORE):
            ws.cell(row=row, column=c, value=val)
        _append_note(ws, row, note_down)
        down_changed += 1
        print(f"{sheet_name} row {row}: D1–D6 {old} -> {new} (sum {sum(old)} -> {sum(new)})")

    overall_changed = 0
    if not no_random_overall:
        if overall_min > overall_max:
            raise ValueError(f"{sheet_name}: overall_min > overall_max")
        for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            if _read_scores(ws, row) is None:
                print(f"{sheet_name} row {row}: skip random Overall (no complete D1–D6)")
                continue
            val = round(random.uniform(overall_min, overall_max), 2)
            ws.cell(row=row, column=COL_OVERALL, value=val)
            _append_note(ws, row, NOTE_OVERALL)
            overall_changed += 1
            print(
                f"{sheet_name} row {row}: Overall (K) = {val} (uniform in [{overall_min}, {overall_max}])"
            )

    return down_changed, overall_changed


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--input", type=Path, default=DEFAULT_IN)
    ap.add_argument("--output", type=Path, default=DEFAULT_OUT)
    ap.add_argument(
        "--seed",
        type=int,
        default=None,
        help="RNG seed for reproducible randomness",
    )
    ap.add_argument(
        "--skip-downscore",
        action="store_true",
        help="Do not apply the −2 random reduction on D1–D6",
    )
    ap.add_argument(
        "--no-random-overall",
        action="store_true",
        help="Leave column K as the workbook formula (no synthetic Overall)",
    )
    ap.add_argument(
        "--overall-min",
        type=float,
        default=3.9,
        help="Prompt A: lower bound for random Overall (K) (default 3.9)",
    )
    ap.add_argument(
        "--overall-max",
        type=float,
        default=4.3,
        help="Prompt A: upper bound for random Overall (K) (default 4.3)",
    )
    ap.add_argument(
        "--prompt-b",
        action="store_true",
        help="Also process sheet 'Prompt B' (same −2 logic; Overall range defaults 4.2–4.67)",
    )
    ap.add_argument(
        "--prompt-b-overall-min",
        type=float,
        default=4.2,
        help="Prompt B: lower bound for random Overall (K)",
    )
    ap.add_argument(
        "--prompt-b-overall-max",
        type=float,
        default=4.67,
        help="Prompt B: upper bound for random Overall (K)",
    )
    args = ap.parse_args()

    if args.overall_min > args.overall_max:
        print("Prompt A: --overall-min must be <= --overall-max", file=sys.stderr)
        return 1
    if args.prompt_b and args.prompt_b_overall_min > args.prompt_b_overall_max:
        print("Prompt B: overall min must be <= max", file=sys.stderr)
        return 1

    src, dst = args.input, args.output
    if not src.is_file():
        print(f"Input not found: {src}", file=sys.stderr)
        return 1

    if args.seed is not None:
        random.seed(args.seed)

    shutil.copy2(src, dst)
    wb = load_workbook(dst)

    jobs: list[tuple[str, float, float]] = [
        ("Prompt A", args.overall_min, args.overall_max),
    ]
    if args.prompt_b:
        jobs.append(
            ("Prompt B", args.prompt_b_overall_min, args.prompt_b_overall_max),
        )

    total_down = total_ov = 0
    for sheet_name, lo, hi in jobs:
        if sheet_name not in wb.sheetnames:
            print(f"Missing sheet {sheet_name!r}", file=sys.stderr)
            return 1
        ws = wb[sheet_name]
        d, o = process_sheet(
            ws,
            sheet_name,
            skip_downscore=args.skip_downscore,
            no_random_overall=args.no_random_overall,
            overall_min=lo,
            overall_max=hi,
        )
        total_down += d
        total_ov += o

    wb.save(dst)
    print(
        f"Wrote: {dst} (sheets: {[j[0] for j in jobs]}; "
        f"downscore rows: {total_down}, random Overall rows: {total_ov})"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
