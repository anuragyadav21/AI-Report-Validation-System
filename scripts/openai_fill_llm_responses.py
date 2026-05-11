#!/usr/bin/env python3
"""Fill LLM response column (D) via OpenAI Chat Completions for each Prompt + Scenario (C).

By default, rows that already have text in column D are skipped so reruns never wipe answers.
Use --overwrite to replace existing LLM text.
"""
from __future__ import annotations

import argparse
import sys
import time
from pathlib import Path

from openai import OpenAI
from openpyxl import load_workbook

from experiment_constants import (
    COL_LLM_RESPONSE,
    COL_PROMPT_SCENARIO,
    FIRST_DATA_ROW,
    LAST_DATA_ROW,
    PROMPT_SHEETS,
)
from openai_env import get_openai_api_key

ROOT = Path(__file__).resolve().parents[1]
DEFAULT_XLSX = ROOT / "prompt-worksheet.xlsx"


def chat_once(
    client: OpenAI,
    *,
    model: str,
    user_prompt: str,
    max_tokens: int,
) -> str:
    resp = client.chat.completions.create(
        model=model,
        messages=[{"role": "user", "content": user_prompt}],
        max_tokens=max_tokens,
        temperature=0.2,
    )
    choice = resp.choices[0]
    content = choice.message.content
    if not content:
        return ""
    return content


def main() -> int:
    p = argparse.ArgumentParser(description=__doc__)
    p.add_argument("--workbook", type=Path, default=DEFAULT_XLSX)
    p.add_argument(
        "--model",
        default="gpt-4o-mini",
        help="OpenAI chat model id (e.g. gpt-4o-mini, gpt-4o)",
    )
    p.add_argument("--max-tokens", type=int, default=8192)
    p.add_argument("--timeout", type=float, default=600.0)
    p.add_argument(
        "--overwrite",
        action="store_true",
        help="Replace column D even when it already has text (default: skip non-empty D)",
    )
    p.add_argument("--dry-run", action="store_true")
    args = p.parse_args()

    api_key = get_openai_api_key(required=not args.dry_run)

    path: Path = args.workbook
    if not path.is_file():
        print(f"Workbook not found: {path}", file=sys.stderr)
        return 1

    client = (
        OpenAI(api_key=api_key, timeout=args.timeout) if not args.dry_run else None
    )

    wb = load_workbook(path)
    for sheet_name in PROMPT_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"Missing sheet {sheet_name!r}; have {wb.sheetnames}", file=sys.stderr)
            return 1

    total = 0
    for sheet_name in PROMPT_SHEETS:
        ws = wb[sheet_name]
        for row in range(FIRST_DATA_ROW, LAST_DATA_ROW + 1):
            if not args.overwrite:
                existing = ws.cell(row=row, column=COL_LLM_RESPONSE).value
                if existing is not None and str(existing).strip():
                    print(f"{sheet_name} row {row}: skip (D already filled; use --overwrite to replace)")
                    continue

            prompt_cell = ws.cell(row=row, column=COL_PROMPT_SCENARIO)
            prompt = prompt_cell.value
            if not prompt or not str(prompt).strip():
                print(f"{sheet_name}!C{row}: empty prompt, skip")
                continue
            prompt = str(prompt)
            total += 1
            print(
                f"[{total}] {sheet_name} row {row} -> OpenAI ({args.model}) ...",
                flush=True,
            )
            if args.dry_run:
                print(f"  (dry-run, {len(prompt)} chars)")
                continue

            assert client is not None
            for attempt in range(3):
                try:
                    text = chat_once(
                        client,
                        model=args.model,
                        user_prompt=prompt,
                        max_tokens=args.max_tokens,
                    )
                    break
                except Exception as e:
                    wait = 2 ** attempt
                    print(f"  attempt {attempt + 1} failed: {e}; retry in {wait}s", flush=True)
                    if attempt == 2:
                        print(f"  giving up on {sheet_name} row {row}", file=sys.stderr)
                        return 1
                    time.sleep(wait)
            else:
                return 1

            ws.cell(row=row, column=COL_LLM_RESPONSE, value=text)
            wb.save(path)
            print(f"  saved D{row} ({len(text)} chars)", flush=True)

    if args.dry_run:
        print(f"Dry-run complete; would run {total} API calls.")
    else:
        print(f"Done. Wrote {total} responses to {path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
